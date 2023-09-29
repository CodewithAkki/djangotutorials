from django.shortcuts import render
from traditional.models import StudentData
from django.http import HttpResponse
import time
from openpyxl import load_workbook
import csv

def GiveTemplate(request):
    if request.method == "POST":
        start = time.time()
        try:
            file = request.FILES['path']
            match file.name.split('.')[1]:
                case 'csv':
                    data = []
                    with open('./'+str(file.name)) as csvfile:
                        reader = csv.DictReader(csvfile)
                        for row in reader:
                            data.append(StudentData(
                                name=row['first_name'],
                                email=row['email'],
                                Designation=row['Designation']
                            ))

                    StudentData.objects.bulk_create(data)
                case 'xlsx':
                    wb = load_workbook(file, read_only=True)
                    sheet = wb.active  # Assuming the first sheet is the one you want to read
                    data = []
                    for row in sheet.iter_rows(min_row=2, values_only=True):
                        name, email, designation = row
                        data.append(StudentData(name=name, email=email, Designation=designation))
                    StudentData.objects.bulk_create(data)
            
            end = time.time()
            print('Time required to run:', (end - start) * 1000, "ms")
            return HttpResponse("Data inserted successfully.")

        except Exception as e:
            return HttpResponse(f"An error occurred: {str(e)}")

    if request.method == "GET":
        return render(request, "index.html")    